# -*- coding: utf-8 -*-
import os
os.environ.setdefault("PYTHONIOENCODING", "utf-8")

import warnings
warnings.filterwarnings(
    "ignore",
    message="Workbook contains no default style, apply openpyxl's default",
    module="openpyxl.styles.stylesheet"
)

import io
import re
import streamlit as st
import pandas as pd
from openpyxl import load_workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter

# Optional OpenAI import
try:
    from openai import OpenAI
except ImportError:
    OpenAI = None

# RapidFuzz for fuzzy matching
try:
    from rapidfuzz import process, fuzz
except ImportError:
    process = None
    fuzz = None

# Streamlit app configuration
st.set_page_config(page_title="Data Combiner & Vendor Tagger", layout="wide")
st.title("📊 Data Combiner & Vendor Tagger")

# --- Sidebar configuration ---
st.sidebar.header("Configuration")
# Optional OpenAI API key
gpt_api_key = st.sidebar.text_input(
    "OpenAI API Key (optional)", type="password",
    help="Provide OpenAI API key to enable GPT-based fallback"
)
# Determine GPT availability
gpt_available = bool(gpt_api_key and OpenAI)
if gpt_available:
    client = OpenAI(api_key=gpt_api_key)
    st.sidebar.success("🔐 GPT fallback enabled")
else:
    client = None
    st.sidebar.info("🤖 Using RapidFuzz only")

# Matching threshold and output filename
threshold = st.sidebar.slider("Fuzzy threshold (0–100)", 0, 100, 80)
output_filename = st.sidebar.text_input("Output filename", "combined.xlsx")

# --- File uploads ---
st.subheader("1. Upload Data Files (.csv, .xlsx)")
uploaded_files = st.file_uploader(
    "Select data files:", type=["csv", "xlsx"], accept_multiple_files=True
)
if uploaded_files:
    st.markdown(f"**{len(uploaded_files)} files selected**")

st.subheader("2. Upload Vendor Master List (.xlsx)")
vendor_file = st.file_uploader(
    "Single-column vendor list:", type=["xlsx"], key="vendor"
)

# --- Run & process ---
if st.button("Run & Download"):
    try:
        # Validate inputs
        if not uploaded_files:
            st.error("Upload at least one data file.")
            st.stop()
        if not vendor_file:
            st.error("Upload the vendor master list.")
            st.stop()

        # Load vendor master list
        df_vendors = pd.read_excel(vendor_file, dtype=str)
        if df_vendors.shape[1] != 1:
            st.error("Vendor list must have exactly one column.")
            st.stop()
        master = df_vendors.iloc[:, 0].dropna().unique().tolist()

        # GPT matching helper
        def gpt_match(name):
            prompt = (
                f"Given the sheet name '{name}' and vendor list {master}, "
                "pick exactly the vendor that best matches or return empty string."
            )
            try:
                resp = client.chat.completions.create(
                    model="gpt-3.5-turbo",
                    messages=[
                        {"role": "system", "content": "You are a helpful assistant."},
                        {"role": "user", "content": prompt}
                    ],
                    temperature=0.0,
                    max_tokens=16
                )
                choice = resp.choices[0].message.content.strip()
            except Exception:
                return ""
            return choice if choice in master else ""

        # Setup progress UI
        total_steps = len(uploaded_files) * 2
        progress = st.progress(0)
        status = st.empty()
        step = 0

        # Build initial workbook in-memory
        buf = io.BytesIO()
        with pd.ExcelWriter(buf, engine="openpyxl") as writer:
            for f in uploaded_files:
                step += 1
                status.text(f"Ingesting ({step}/{total_steps}): {f.name}")
                df = (
                    pd.read_csv(f, dtype=str)
                    if f.name.lower().endswith(".csv")
                    else pd.read_excel(f, dtype=str)
                )
                df.insert(0, "Vendor", "")
                sheet_name = f.name.rsplit('.', 1)[0][:31]
                df.to_excel(writer, sheet_name=sheet_name, index=False)
                progress.progress(step / total_steps)
        buf.seek(0)

        # Load workbook and style + match
        wb = load_workbook(buf)
        for ws in wb.worksheets:
            step += 1
            idx = step - len(uploaded_files)
            status.text(f"Styling & matching ({idx}/{len(uploaded_files)}): {ws.title}")
            max_row, max_col = ws.max_row, ws.max_column
            if max_row < 2 or max_col < 1:
                progress.progress(step / total_steps)
                continue
            last_col = get_column_letter(max_col)
            ref = f"A1:{last_col}{max_row}"
            safe = re.sub(r'[^A-Za-z0-9_]', '_', ws.title)
            tbl = Table(displayName=f"tbl_{safe}", ref=ref)
            tbl.tableStyleInfo = TableStyleInfo(
                name="TableStyleMedium9",
                showFirstColumn=False,
                showLastColumn=False,
                showRowStripes=True,
                showColumnStripes=False
            )
            ws.add_table(tbl)

            # Fuzzy matching
            vendor = ""
            if process and fuzz:
                m = process.extractOne(ws.title, master, scorer=fuzz.partial_ratio)
                if m and m[1] >= threshold:
                    vendor = m[0]
            # GPT fallback
            if not vendor and gpt_available:
                vendor = gpt_match(ws.title)

            for row in ws.iter_rows(min_row=2, max_row=max_row, min_col=1, max_col=1):
                row[0].value = vendor or ""
            progress.progress(step / total_steps)

        status.empty()
        progress.empty()

        # Save and download
        out_buf = io.BytesIO()
        wb.save(out_buf)
        out_buf.seek(0)
        st.success("✅ Processing complete!")
        st.download_button(
            "Download Combined Workbook", out_buf,
            file_name=output_filename,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
    except Exception as e:
        st.error(f"An error occurred: {e}")
        st.exception(e)
